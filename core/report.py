#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""报告生成：CSV / Excel"""
import os
import csv
from datetime import datetime
from pathlib import Path


def generate_csv_report(results, output_dir):
    """生成 CSV 分类报告
    results: {category: [(img_path, reason), ...]}
    """
    report_path = Path(output_dir) / f"分类报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    with open(report_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "文件名", "分类结果", "分类原因", "原路径"])
        row = 1
        for category, files in results.items():
            for img_path, reason in files:
                writer.writerow([row, os.path.basename(img_path), category, reason, img_path])
                row += 1
    return str(report_path)


def generate_excel_report(results, output_dir):
    """生成 Excel 分类报告（需要 openpyxl）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # 未安装 openpyxl 时回退到 CSV
        return generate_csv_report(results, output_dir)

    report_path = Path(output_dir) / f"分类报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分类结果"

    headers = ["序号", "文件名", "分类结果", "分类原因", "原路径"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 1
    for category, files in results.items():
        for img_path, reason in files:
            ws.append([row, os.path.basename(img_path), category, reason, img_path])
            row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 60

    wb.save(report_path)
    return str(report_path)
